import collections.abc
import datetime
import io
import json
import math
import os
import re
import urllib.error
import urllib.parse
from tempfile import NamedTemporaryFile
from wsgiref.util import FileWrapper

import numpy as np
import openpyxl
import pandas as pd
from auditlog.models import LogEntry
from django.contrib.contenttypes.models import ContentType
from django.core.files.storage import default_storage
from django.db import transaction
from django.db.models import Q
from django.db.models.deletion import ProtectedError
from django.db.utils import IntegrityError, NotSupportedError
from django.http import FileResponse, Http404, HttpResponse, JsonResponse
from django.shortcuts import redirect
from rest_framework import status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.serializers import (
    BooleanField,
    CharField,
    FileField,
    ModelSerializer,
    PrimaryKeyRelatedField,
    SerializerMethodField,
    StringRelatedField,
    ValidationError,
)
from rest_framework.views import APIView
from rest_framework.viewsets import ModelViewSet

from antigenapi.bioinformatics.blast import get_db_fasta, run_blastp
from antigenapi.bioinformatics.imgt import (
    AIRR_IMPORTANT_COLUMNS,
    as_fasta_files,
    load_sequences,
    read_airr_file,
    run_vquest,
)
from antigenapi.models import (
    Antigen,
    Cohort,
    ElisaPlate,
    ElisaWell,
    Library,
    Llama,
    Nanobody,
    PlateLocations,
    Project,
    SequencingRun,
    SequencingRunResults,
)
from antigenapi.utils.uniprot import get_protein

from .parsers import parse_elisa_file


# Audit logs #
class AuditLogSerializer(ModelSerializer):
    """A serializer for object audit logs."""

    actor_username = StringRelatedField(source="actor.username", read_only=True)
    actor_email = StringRelatedField(source="actor.email", read_only=True)

    class Meta:  # noqa: D106
        model = LogEntry
        fields = [
            "timestamp",
            "object_id",
            "action",
            "changes_dict",
            "actor",
            "actor_username",
            "actor_email",
        ]


# Mixins #
class DeleteProtectionMixin(object):
    """Show helpful error when delete protection is in place (on_delete=PROTECT)."""

    def destroy(self, request, *args, **kwargs):
        """Override destroy method to catch Django's ProtectedError, return HTTP 400."""
        try:
            return super().destroy(request, *args, **kwargs)
        except ProtectedError as protected_error:
            protected_elements = set()
            for obj in protected_error.protected_objects:
                if isinstance(obj, ElisaWell):
                    protected_elements.add(str(obj.plate))
                else:
                    protected_elements.add(str(obj))
            protected_elements = tuple(protected_elements)
            msg = f"DeleteProtection: Object in use by {protected_elements[0]}"
            if len(protected_elements) > 1:
                msg += f" and {len(protected_elements) - 1} other object"
                if len(protected_elements) > 2:
                    msg += "s"
            response_data = {"message": msg, "protected_elements": protected_elements}
            return Response(data=response_data, status=status.HTTP_400_BAD_REQUEST)


class AuditLogMixin(object):
    """Allow fetching audit logs on individual objects."""

    @action(detail=True, methods=["GET"], name="Get audit log")
    def auditlog(self, request, pk):
        """Get audit logs for an object."""
        queryset = LogEntry.objects.filter(
            content_type=ContentType.objects.get_for_model(self.queryset.model),
            object_id=pk,
        ).select_related("actor")

        serializer = AuditLogSerializer(queryset, many=True)
        return Response(serializer.data)


# Projects #
class ProjectSerializer(ModelSerializer):
    """A serializer for project data which serializes all internal fields."""

    added_by = StringRelatedField()

    class Meta:  # noqa: D106
        model = Project
        fields = "__all__"
        read_only_fields = ["added_by", "added_date"]


class ProjectViewSet(AuditLogMixin, DeleteProtectionMixin, ModelViewSet):
    """A view set displaying all recorded projects."""

    queryset = Project.objects.all().select_related("added_by").order_by("short_title")
    serializer_class = ProjectSerializer

    def perform_create(self, serializer):
        """Overload the perform_create method."""
        serializer.save(added_by=self.request.user)


# Llamas #
class LlamaSerializer(ModelSerializer):
    """A serializer for llamas."""

    added_by = StringRelatedField()

    class Meta:  # noqa: D106
        model = Llama
        fields = "__all__"
        read_only_fields = ["added_by", "added_date"]


class LlamaViewSet(AuditLogMixin, DeleteProtectionMixin, ModelViewSet):
    """A view set for llamas."""

    queryset = Llama.objects.all().select_related("added_by").order_by("name")
    serializer_class = LlamaSerializer

    def perform_create(self, serializer):  # noqa: D102
        serializer.save(added_by=self.request.user)


# Library #
class LibrarySerializer(ModelSerializer):
    """A serializer for libraries."""

    added_by = StringRelatedField()
    # llama_name = CharField(source='llama.name', read_only=True)
    cohort_cohort_num = SerializerMethodField(
        read_only=True
    )  # CharField(source="cohort.cohort_num", read_only=True)
    cohort_is_naive = BooleanField(source="cohort.is_naive", read_only=True)
    cohort_cohort_num_prefixed = SerializerMethodField(read_only=True)
    project_short_title = CharField(source="project.short_title", read_only=True)
    library_num = CharField(read_only=True)

    class Meta:  # noqa: D106
        model = Library
        fields = "__all__"
        read_only_fields = ["added_by", "added_date"]

    def get_cohort_cohort_num(self, obj):
        """Get the cohort number with sublibrary, if present."""
        return f"{obj.cohort.cohort_num}{obj.sublibrary or ''}"

    def get_cohort_cohort_num_prefixed(self, obj):
        """Get the cohort number with sublibrary and N (naive) prefix."""
        return f"{obj.cohort.cohort_num_prefixed()}{obj.sublibrary or ''}"


class LibraryViewSet(AuditLogMixin, DeleteProtectionMixin, ModelViewSet):
    """A view set for libraries."""

    queryset = (
        Library.objects.all()
        .select_related("cohort")
        .select_related("project")
        .select_related("added_by")
        .order_by("cohort__is_naive", "cohort__cohort_num")
    )
    serializer_class = LibrarySerializer
    filterset_fields = ("project",)

    def perform_create(self, serializer):  # noqa: D102
        try:
            serializer.save(added_by=self.request.user)
        except IntegrityError as e:
            raise ValidationError({"non_field_errors": [str(e)]})


class AntigenSerializer(ModelSerializer):
    """A serializer for antigen data.

    A serializer for antigen data which serializes all internal fields, and includes the
    serialzed related local or UniProt antigen data and provides a set of elisa well
    which reference it.
    """

    added_by = StringRelatedField()
    elisa_plates = SerializerMethodField()
    sequencing_runs = SerializerMethodField()

    class Meta:  # noqa: D106
        model = Antigen
        fields = "__all__"
        read_only_fields = ["added_by", "added_date"]

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)

        request = self.context.get("request")
        if request and not request.parser_context.get("kwargs", {}).get("pk"):
            # If it's a list view, remove these fields
            self.fields.pop("elisa_plates", None)
            self.fields.pop("sequencing_runs", None)

    def get_elisa_plates(self, obj):
        """Get ELISA plates using this antigen in single-object requests."""
        # Only include related data if this is a single-object request
        request = self.context.get("request")
        if request and request.parser_context.get("kwargs", {}).get("pk"):
            return ElisaPlateWithoutWellsSerializer(
                ElisaPlate.objects.filter(elisawell__antigen=obj.pk).distinct(),
                many=True,
            ).data
        return None  # Omit in list views

    def get_sequencing_runs(self, obj):
        """Get sequencing runs using this antigen in single-object requests."""
        # Only include related data if this is a single-object request
        request = self.context.get("request")
        if request and request.parser_context.get("kwargs", {}).get("pk"):
            elisa_plate_ids = (
                ElisaPlate.objects.filter(elisawell__antigen=obj.pk)
                .values_list("id", flat=True)
                .distinct()
            )

            # Handle case where no ELISAs available
            if not elisa_plate_ids:
                return SequencingRunSerializer(many=True).data

            # Try PostgreSQL JSONField filtering
            # Build a Q object to check if any given elisa_plate is inside the JSONField
            query = Q()
            for plate_id in elisa_plate_ids:
                query |= Q(plate_thresholds__contains=[{"elisa_plate": plate_id}])

            sequencing_runs = SequencingRun.objects.filter(query)

            try:
                return SequencingRunShortSerializer(sequencing_runs, many=True).data
            except NotSupportedError:
                # Fallback for SQLite: Manual filtering in Python
                sequencing_runs = [
                    run
                    for run in SequencingRun.objects.all()
                    if any(
                        entry.get("elisa_plate") in elisa_plate_ids
                        for entry in run.plate_thresholds
                    )
                ]
                return SequencingRunShortSerializer(sequencing_runs, many=True).data
        return None

    def validate(self, data):
        """Check the antigen is a valid uniprot ID."""
        if data.get("uniprot_id"):
            try:
                protein_data = get_protein(data["uniprot_id"])
            except urllib.error.HTTPError as e:
                if e.code == 400:
                    raise ValidationError(
                        {"uniprot_id": "Couldn't validate this UniProt ID (code 400)"}
                    )
                elif e.code == 500:
                    raise ValidationError(
                        {"uniprot_id": "Couldn't validate this UniProt ID (code 500)"}
                    )
                else:
                    raise
            if not data.get("sequence") or data.get("sequence").strip() == "":
                data["sequence"] = protein_data["sequence"]
            if data.get("molecular_mass") is None:
                data["molecular_mass"] = protein_data["molecular_mass"]
            if not data.get("long_name") or data.get("long_name").strip() == "":
                data["long_name"] = protein_data["protein_name"]

        return data


class AntigenViewSet(AuditLogMixin, DeleteProtectionMixin, ModelViewSet):
    """A view set displaying all recorded antigens."""

    queryset = Antigen.objects.all().select_related("added_by").order_by("short_name")
    serializer_class = AntigenSerializer

    def perform_create(self, serializer):  # noqa: D102
        serializer.save(added_by=self.request.user)


# Cohort #
class CohortSerializer(ModelSerializer):
    """A serializer for cohorts."""

    added_by = StringRelatedField()
    llama_name = CharField(source="llama.name", read_only=True)
    antigen_details = AntigenSerializer(source="antigens", many=True, read_only=True)
    # project_short_title = CharField(source='project.short_title', read_only=True)
    cohort_num_prefixed = CharField(read_only=True)

    class Meta:  # noqa: D106
        model = Cohort
        fields = "__all__"
        read_only_fields = ["added_by", "added_date"]


class CohortViewSet(AuditLogMixin, DeleteProtectionMixin, ModelViewSet):
    """A view set for cohorts."""

    queryset = (
        Cohort.objects.all().select_related("llama").order_by("is_naive", "cohort_num")
    )
    serializer_class = CohortSerializer
    filterset_fields = ("llama", "cohort_num")

    def perform_create(self, serializer):  # noqa: D102
        serializer.save(added_by=self.request.user)


# ELISA plates #
class NestedElisaWellSerializer(ModelSerializer):
    """A serializer for elisa wells."""

    optical_density = SerializerMethodField()

    class Meta:  # noqa: D106
        model = ElisaWell
        exclude = ("id", "plate")

    def get_optical_density(self, obj):
        """Get optical density - convert NaN to None for JSON encoding."""
        if obj.optical_density is not None and not math.isnan(obj.optical_density):
            return obj.optical_density
        return None


class ElisaPlateWithoutWellsSerializer(ModelSerializer):
    """A serializer for elisa plates without well data included.

    A serializer for elisa plates which serializes all internal fields.
    """

    project_short_title = CharField(
        source="library.project.short_title", read_only=True
    )
    library_cohort_cohort_num = CharField(
        source="library.cohort.cohort_num", required=False
    )
    library_cohort_cohort_num_prefixed = CharField(
        source="library.cohort.cohort_num_prefixed", required=False
    )
    library_cohort_is_naive = BooleanField(
        source="library.cohort.is_naive", read_only=True
    )
    library_library_num = CharField(source="library.library_num", read_only=True)
    added_by = StringRelatedField()
    antigen = PrimaryKeyRelatedField(queryset=Antigen.objects.all(), write_only=True)
    read_only_fields = [
        "library_cohort_cohort_num",
        "elisawell_set",
        "added_by",
        "added_date",
    ]
    plate_file = FileField(use_url=False)

    class Meta:  # noqa: D106
        model = ElisaPlate
        fields = "__all__"


class ElisaPlateSerializer(ElisaPlateWithoutWellsSerializer):
    elisawell_set = NestedElisaWellSerializer(many=True, required=False)

    """A serializer for elisa plates included well data.
    """

    def validate(self, data):
        """Validate plate (load and parse file)."""
        if "plate_file" in data:
            try:
                data["elisawell_set"] = parse_elisa_file(data["plate_file"])
            except Exception as e:
                raise ValidationError({"plate_file": e})
        return data

    @staticmethod
    def _create_wells(plate, antigen, well_set):
        ElisaWell.objects.bulk_create(
            ElisaWell(plate=plate, optical_density=od, location=loc, antigen=antigen)
            for (od, loc) in zip(well_set, PlateLocations)
        )

    @transaction.atomic
    def create(self, validated_data):
        """Create plate. For now, every well shares the same antigen."""
        antigen = validated_data.pop("antigen")
        well_set = validated_data.pop("elisawell_set")
        validated_data["plate_file"].seek(0)
        plate = super(ElisaPlateSerializer, self).create(validated_data)
        self._create_wells(plate, antigen, well_set)
        return plate

    @transaction.atomic
    def update(self, instance, validated_data):
        """Update plate. For now, every well shares the same antigen."""
        antigen = validated_data["antigen"]
        if "elisawell_set" in validated_data:
            well_set = validated_data.pop("elisawell_set")
        else:
            well_set = None
        if "plate_file" in validated_data:
            try:
                validated_data["plate_file"].seek(0)
            except ValueError:
                # File has already been read during create, so skip it
                validated_data.pop("plate_file")
                well_set = None
        plate = super(ElisaPlateSerializer, self).update(instance, validated_data)
        if well_set is not None:
            # Faster, fewer DB queries to bulk delete & re-insert
            # than conditionally update
            ElisaWell.objects.filter(plate=plate).delete()
            self._create_wells(plate, antigen, well_set)
        else:
            ElisaWell.objects.filter(plate=plate).update(antigen=antigen)
        return instance


def _wells_to_tsv(wells):
    UPPER_CASE_A = 65
    ROW_LENGTH = 12
    NUM_ROWS = 8

    output = "\t".join([""] + [str(i) for i in range(1, ROW_LENGTH + 1)]) + "\n"
    for row in range(NUM_ROWS):
        start = row * ROW_LENGTH
        row = [chr(UPPER_CASE_A + row)]
        row += [
            str(w) if w is not None else "" for w in wells[start : (start + ROW_LENGTH)]
        ]
        output += "\t".join(row) + "\n"

    return output


class ElisaPlateViewSet(AuditLogMixin, DeleteProtectionMixin, ModelViewSet):
    """A view set displaying all recorded elisa plates."""

    queryset = (
        ElisaPlate.objects.all()
        .select_related("library__cohort")
        .select_related("library__project")
        .select_related("added_by")
        .prefetch_related("elisawell_set")
        .order_by("-added_date")
    )
    serializer_class = ElisaPlateSerializer
    filterset_fields = ("library", "library__cohort")

    def perform_create(self, serializer):  # noqa: D102
        serializer.save(added_by=self.request.user)

    @action(
        detail=True,
        methods=["GET"],
        name="Download ELISA plate as TSV.",
        url_path="tsv",
    )
    def download_elisa_tsv(self, request, pk):
        """Download ELISA plate as .tsv file."""
        wells = list(
            ElisaWell.objects.filter(
                plate_id=pk,
            )
            .order_by("location")
            .values_list("optical_density", flat=True)
        )

        output = _wells_to_tsv(wells)

        response = HttpResponse(output, content_type="text/tab-separated-values")

        response["Content-Disposition"] = f'attachment; filename="elisa_plate_{pk}.tsv"'

        return response


class ElisaWellInlineSerializer(ModelSerializer):
    """A serializer to represent elisa wells by plate id and location."""

    class Meta:  # noqa: D106
        model = ElisaWell
        fields = ("plate", "location")


# Nanobodies #
class NanobodySerializer(ModelSerializer):
    """A serializer for nanobodies."""

    added_by = StringRelatedField()

    class Meta:  # noqa: D106
        model = Nanobody
        fields = "__all__"
        read_only_fields = ["seqruns", "added_by", "added_date"]


class NanobodyViewSet(AuditLogMixin, DeleteProtectionMixin, ModelViewSet):
    """A view set for nanobodies."""

    queryset = Nanobody.objects.all().select_related("added_by").order_by("name")
    serializer_class = NanobodySerializer

    def perform_create(self, serializer):  # noqa: D102
        serializer.save(added_by=self.request.user)


class SequencingRunResultSerializer(ModelSerializer):
    """A serializer for sequencing run results."""

    added_by = StringRelatedField()

    class Meta:  # noqa: D106
        model = SequencingRunResults
        fields = ("seq", "well_pos_offset", "added_by", "added_date")


class SequencingRunSerializer(ModelSerializer):
    """A serializer for sequencing runs."""

    added_by = StringRelatedField()
    sequencingrunresults_set = SequencingRunResultSerializer(many=True, required=False)

    def validate_wells(self, data):
        """Check JSONField for wells is valid."""
        for idx, well in enumerate(data):
            if len(well.keys()) != 3:
                raise ValidationError(f"Extraneous keys in well {idx}")
            if "elisa_well" not in well:
                raise ValidationError(f"Missing elisa_well in well {idx}")

            if "plate" not in well:
                raise ValidationError(f"Missing plate in well {idx}")
            if not isinstance(well["plate"], int):
                raise ValidationError(f"Well {idx}'s plate is not an integer")
            if well["plate"] != (idx // 96):
                raise ValidationError(
                    f"Well {idx}'s plate should be "
                    f"{(idx // 96)} (found: {well['plate']})"
                )

            if "location" not in well:
                raise ValidationError(f"Missing location in well {idx}")
            if not isinstance(well["location"], int):
                raise ValidationError(f"Well {idx}'s location is not an integer")
            if well["location"] < 1 or well["location"] > 96:
                raise ValidationError(
                    f"Well {idx}'s location must be between 1 and 96 inclusive"
                )

            if not isinstance(well["elisa_well"], collections.abc.Mapping):
                raise ValidationError(f"elisa_well in well {idx} is not an object")
            if len(well["elisa_well"].keys()) != 2:
                raise ValidationError(f"Extraneous keys in well {idx}'s elisa_well")
            if "plate" not in well["elisa_well"]:
                raise ValidationError(f"Missing plate in well {idx}'s elisa_well")
            if not isinstance(well["elisa_well"]["plate"], int):
                raise ValidationError(
                    f"Well {idx}'s elisa_well's plate is not an integer"
                )
            if "location" not in well["elisa_well"]:
                raise ValidationError(f"Missing location in well {idx}'s elisa_well")
            if not isinstance(well["elisa_well"]["location"], int):
                raise ValidationError(
                    f"Well {idx}'s elisa_well's location is not an integer"
                )
            if (
                well["elisa_well"]["location"] < 1
                or well["elisa_well"]["location"] > 96
            ):
                raise ValidationError(
                    f"Well {idx}'s elisa_well's location must be "
                    "between 1 and 96 inclusive"
                )

            # TODO: Check elisa_plate is valid plate
            # TODO: Check all locations are unique on a plate

        # Sort by plate and location
        data = sorted(data, key=lambda well: (well["plate"], well["location"]))

        return data

    def validate_plate_thresholds(self, data):
        """Check JSONField for plate_thresholds is valid."""
        for idx, thr in enumerate(data):
            if len(thr.keys()) != 2:
                raise ValidationError(f"Extraneous keys in plate_threshold {idx}")
            if "optical_density_threshold" not in thr:
                raise ValidationError(
                    f"Optical density threshold missing in plate_threshold {idx}"
                )
            if not isinstance(thr["optical_density_threshold"], (int, float)):
                raise ValidationError(
                    f"Plate threshold {idx}'s optical_density_threshold "
                    "is not an integer or float"
                )
            if thr["optical_density_threshold"] < 0:
                raise ValidationError(
                    f"Plate threshold {idx}'s optical_density_threshold "
                    "must be at least 0"
                )
            if "elisa_plate" not in thr:
                raise ValidationError(f"Elisa plate missing in plate_threshold {idx}")
            if not isinstance(thr["elisa_plate"], int):
                raise ValidationError(
                    f"Plate threshold {idx}'s elisa_plate is not an integer"
                )

            # TODO: Check elisa_plate is valid plate
            # TODO: Check elisa_plate is unique within list
            # TODO: Check threshold is present for every plate listed in elisa_well

        return data

    class Meta:  # noqa: D106
        model = SequencingRun
        fields = "__all__"
        read_only_fields = ["added_by", "added_date"]


class SequencingRunShortSerializer(SequencingRunSerializer):
    """Sequencing run serializer which excludes plate/well information."""

    class Meta:  # noqa: D106
        model = SequencingRun
        exclude = ("plate_thresholds", "wells")
        read_only_fields = ["added_by", "added_date"]


def _extract_well(well):
    well = well.upper()
    try:
        # Match A1-H12, including A01 etc.
        well_match = re.search("[A-H]((1[0-2])|(0?[1-9]))$", well).group(0)
    except AttributeError:
        raise ValueError("Unable to extract well name from filename")
    # Remove zero-padding if present
    if well_match[1] == "0":
        well_match = well_match[0] + well_match[2]
    return well_match


class SequencingRunViewSet(AuditLogMixin, DeleteProtectionMixin, ModelViewSet):
    """A view set for sequencing runs."""

    queryset = SequencingRun.objects.all().order_by("-added_date")
    serializer_class = SequencingRunSerializer

    def perform_create(self, serializer):  # noqa: D102
        serializer.save(added_by=self.request.user)

    @action(
        detail=True,
        methods=["GET"],
        name="Download sequencing run submission file (xlsx).",
        url_path="submissionfile/(?P<submission_idx>[0-9]+)",
    )
    def download_submission_xlsx(self, request, pk, submission_idx):
        """Download sequencing run submission file (xlsx)."""
        # Load the Excel template
        fn = os.path.join(
            os.path.dirname(os.path.realpath(__file__)),
            "files",
            "sequencing-submission-form-v1.xlsx",
        )
        oxl = openpyxl.load_workbook(fn)

        # Get first worksheet
        ws = oxl.worksheets[0]

        # Get sequencing run object
        sr = self.get_object()
        well_dict = {}
        elisa_plates_to_load = set()
        submission_idx = int(submission_idx)
        for well in sr.wells:
            if well["plate"] != submission_idx:
                continue
            well_dict[PlateLocations.labels[well["location"] - 1]] = well["elisa_well"]
            elisa_plates_to_load.add(well["elisa_well"]["plate"])

        if not well_dict:
            raise Http404

        elisa_wells = {
            (ew.plate_id, ew.location): ew
            for ew in ElisaWell.objects.filter(plate__pk__in=elisa_plates_to_load)
            .select_related("plate")
            .select_related("antigen")
        }

        # Make modifications
        for row in range(3, 100):  # 96 wells
            try:
                # Sample name
                well = well_dict[ws[f"A{row}"].value]
            except KeyError:
                continue

            elisa_well = elisa_wells[(well["plate"], well["location"])]
            ws[f"B{row}"] = (
                f"{elisa_well.antigen.short_name}_"
                f"EP{elisa_well.plate_id}_"
                f"{PlateLocations.labels[elisa_well.location - 1]}"
            )
            # Own primer name
            ws[f"D{row}"] = "PHD_SEQ_FWD"
            ws[f"F{row}"] = "Yes"

        # Save to temp file
        with NamedTemporaryFile() as tmp:
            oxl.save(tmp.name)
            response = HttpResponse(
                FileWrapper(tmp),
                content_type="application/vnd.openxmlformats-officedocument"
                ".spreadsheetml.sheet",
            )

        response["Content-Disposition"] = (
            f'attachment; filename="sequencing-submission-form-'
            f'sr{pk}_{submission_idx}.xlsx"'
        )
        return response

    @action(
        detail=True,
        methods=["GET"],
        name="Download sequencing run submission file (xlsx).",
        url_path="submissionfile/(?P<submission_idx>[0-9]+)/tsv",
    )
    def download_sequencing_plate_tsv(self, request, pk, submission_idx):
        """Download sequencing run plate layout as .tsv file."""
        try:
            sr = SequencingRun.objects.get(id=int(pk))
        except SequencingRunResults.DoesNotExist:
            raise Http404

        wells = {
            w["location"]: w for w in sr.wells if w["plate"] == int(submission_idx)
        }
        plate_ids = [w["elisa_well"]["plate"] for w in wells.values()]
        elisa_wells = {
            (ew.plate_id, ew.location): ew
            for ew in ElisaWell.objects.filter(plate_id__in=plate_ids)
        }

        well_dat = []
        for i in range(1, 97):
            try:
                well = wells[i]
            except KeyError:
                well_dat.append("")
                continue

            elisa_well = elisa_wells[
                (well["elisa_well"]["plate"], well["elisa_well"]["location"])
            ]

            well_dat.append(
                f"{elisa_well.plate_id}:"
                f"{PlateLocations.labels[elisa_well.location - 1]} "
                f"[{elisa_well.antigen}]"
            )

        output = _wells_to_tsv(well_dat)

        response = HttpResponse(output, content_type="text/tab-separated-values")

        fn = f"sequencing_run_{pk}_plate_{submission_idx}.tsv"
        response["Content-Disposition"] = f'attachment; filename="{fn}"'

        return response

    @action(
        detail=True,
        methods=["GET", "PUT"],
        name="Upload/download sequencing run results file (.zip).",
        url_path="resultsfile/(?P<submission_idx>[0-9]+)",
    )
    def sequencing_run_results(self, request, pk, submission_idx):
        """Upload sequencing run results file (.zip)."""
        if request.method == "GET":
            return self.download_sequencing_run_results(request, pk, submission_idx)

        results_file = request.data["file"]

        # TODO: Validate results file in more detail
        if not results_file.name.endswith(".zip"):
            raise ValidationError("file", "Results file should be a .zip file")

        # Validate the plate number and submission idx (seq) from the URL
        try:
            sr = SequencingRun.objects.get(pk=int(pk))
        except SequencingRun.DoesNotExist:
            raise ValidationError(
                f"Sequencing run {pk} does not exist to attach results"
            )

        # Store well positions
        wells_expected_list = [
            PlateLocations.labels[w["location"] - 1]
            for w in sr.wells
            if w["plate"] == int(submission_idx)
        ]
        if not wells_expected_list:
            raise ValidationError(
                f"Plate index {submission_idx} not found in sequencing run {pk}"
            )

        # Run bioinformatics using .zip file
        # print("Extracting zip file...")
        try:
            seq_data_fh = results_file.temporary_file_path()
        except AttributeError:
            seq_data_fh = results_file.file
        try:
            seq_data = load_sequences(seq_data_fh)
        except ValueError as e:
            raise ValidationError({"file": str(e)})
        if len(seq_data) != len(wells_expected_list):
            raise ValidationError(
                {
                    "file": f"Upload contains data for {len(seq_data)} "
                    f"wells, expected {len(wells_expected_list)}"
                }
            )
        # Validate wells expected vs wells supplied
        try:
            wells_supplied_list = [_extract_well(w) for w in seq_data.keys()]
        except IndexError:
            raise ValidationError(
                {
                    "file": "Unable to parse well names. "
                    "Ensure all .seq filenames end with a well."
                }
            )

        # Check for duplicates in the supplied well names, and error if so
        wells_supplied = set(wells_supplied_list)
        if len(wells_supplied_list) != len(wells_supplied):
            seen = set()
            seen_twice = list(
                set(w for w in wells_supplied_list if w in seen or seen.add(w))
            )

            raise ValidationError(
                {"file": f"Duplicate wells found in supplied list: {seen_twice}"}
            )

        # Check for duplicates in the expected well names, and error if so
        wells_expected = set(wells_expected_list)
        if len(wells_expected_list) != len(wells_expected):
            raise ValueError(
                f"Duplicate wells found in expected list for seq run {pk} "
                f"idx {submission_idx}"
            )

        # Apply an offset in case the wells are shifted
        offset = 0

        wells_supplied_int = [PlateLocations.labels.index(w) for w in wells_supplied]
        wells_expected_int = [PlateLocations.labels.index(w) for w in wells_expected]
        if min(wells_supplied_int) > min(wells_expected_int):
            offset = min(wells_supplied_int) - min(wells_expected_int)
            wells_expected = set(
                [PlateLocations.labels[w + offset] for w in wells_expected_int]
            )

        if wells_expected - wells_supplied:
            raise ValidationError(
                {
                    "file": f"Expected well(s) {wells_expected - wells_supplied} "
                    f"were not found in upload (h.offset: {offset})"
                }
            )
        if wells_supplied - wells_expected:
            raise ValidationError(
                {
                    "file": f"Unexpected well(s) {wells_supplied - wells_expected} "
                    f"were found in upload (h.offset: {offset})"
                }
            )

        # Convert to FASTA in-memory (vquest api handles chunking to
        # max 50 seqs per file)
        fasta_file = as_fasta_files(seq_data, max_file_size=None)[0]

        # Submit to vquest
        # print("Running vquest...")
        vquest_results = run_vquest(fasta_file)

        parameters_file_data = vquest_results["Parameters.txt"]
        vquest_airr_data = vquest_results["vquest_airr.tsv"]

        base_filename = f"SequencingResults_{pk}_{submission_idx}"

        # Link to nanobodies table
        sequences = read_airr_file(
            io.BytesIO(vquest_airr_data.encode()), usecols=("sequence_alignment_aa",)
        )
        sequences = (
            sequences["sequence_alignment_aa"]
            .dropna()
            .str.replace(".", "")
            .replace("*", "X")
            .unique()
        )
        nanobodies = Nanobody.objects.filter(sequence__in=sequences)

        # Create SequencingRunResults object
        srr, _ = SequencingRunResults.objects.update_or_create(
            sequencing_run=SequencingRun.objects.get(pk=int(pk)),
            seq=submission_idx,
            defaults={
                "added_by": request.user,
                "seqres_file": results_file,
                "well_pos_offset": offset,
            },
        )
        srr.nanobodies.set(nanobodies)

        srr.airr_file.save(
            f"{base_filename}_vquestairr.tsv", io.StringIO(vquest_airr_data)
        )
        srr.parameters_file.save(
            f"{base_filename}_vquestparams.txt", io.StringIO(parameters_file_data)
        )

        return JsonResponse(
            SequencingRunSerializer(SequencingRun.objects.get(pk=int(pk))).data
        )

    def download_sequencing_run_results(self, request, pk, submission_idx):
        """Download sequencing run results file (.zip)."""
        try:
            sr = SequencingRunResults.objects.get(
                sequencing_run_id=int(pk), seq=int(submission_idx)
            )
        except SequencingRunResults.DoesNotExist:
            raise Http404

        if default_storage.__class__.__name__ == "S3Storage":
            return redirect(sr.seqres_file.url)

        filename = sr.seqres_file.path
        response = FileResponse(open(filename, "rb"))
        return response

    @action(
        detail=True,
        methods=["GET"],
        name="Download IMGT results file (.airr).",
        url_path="resultsfile/(?P<submission_idx>[0-9]+)/airr",
    )
    def download_sequencing_run_airr(self, request, pk, submission_idx):
        """Download sequencing run results file (.zip)."""
        try:
            sr = SequencingRunResults.objects.get(
                sequencing_run_id=int(pk), seq=int(submission_idx)
            )
        except SequencingRunResults.DoesNotExist:
            raise Http404

        if default_storage.__class__.__name__ == "S3Storage":
            return redirect(sr.airr_file.url)

        filename = sr.airr_file.path
        response = FileResponse(open(filename, "rb"))
        return response

    @action(
        detail=True,
        methods=["GET"],
        name="Get sequencing results.",
        url_path="results",
    )
    def get_sequencing_run_results(self, request, pk):
        """Get sequencing results."""
        results = (
            SequencingRunResults.objects.filter(sequencing_run_id=int(pk))
            .order_by("seq")
            .select_related("sequencing_run")
            .prefetch_related("nanobodies")
        )

        if not results:
            return JsonResponse({"records": []})

        # Nanobody autoname format is:
        # <short antigen name>_<pan conc>_<ELISA well_no>
        # [.<ELISA plate number (index) if >1 plate]_C<cohort><sublibrary>

        # Get ELISA wells as dict for lookup
        elisa_wells_to_seq = {
            (w["elisa_well"]["plate"], w["elisa_well"]["location"]): (
                w["plate"],
                w["location"],
            )
            for w in results[0].sequencing_run.wells
        }
        # Get the ELISA plate IDs seen across this resultset -
        # put in dict for an ordered set
        elisa_plate_idxs = dict.fromkeys(w[0] for w in elisa_wells_to_seq.keys())

        if len(elisa_plate_idxs) == 1:
            # If there's only one ELISA plate in this set,
            # don't include it in the nanobody identifier
            elisa_plate_idxs = {list(elisa_plate_idxs.keys())[0]: ""}
        else:
            # Otherwise, we include a .(1-based index) suffix to disambiguate
            elisa_plate_idxs = {
                id: f".{idx + 1}" for idx, id in enumerate(elisa_plate_idxs.keys())
            }

        # Retrieve nanobody autonames associated with ELISA plates in this result set
        nanobody_autonames_lookup = {
            elisa_wells_to_seq[(ew.plate_id, ew.location)]: f"{ew.antigen.short_name}_"
            f"{ew.plate.pan_round_concentration:g}_"
            f"{PlateLocations.labels[ew.location - 1]}{elisa_plate_idxs[ew.plate_id]}_C"
            + ("N" if ew.plate.library.cohort.is_naive else "")
            + f"{ew.plate.library.cohort.cohort_num}"
            + f"{ew.plate.library.sublibrary or ''}"
            for ew in ElisaWell.objects.filter(
                plate__in=elisa_plate_idxs.keys()
            ).select_related(
                "antigen", "plate", "plate__library", "plate__library__cohort"
            )
            if (ew.plate_id, ew.location) in elisa_wells_to_seq.keys()
        }

        csvs = []
        for r in results:
            airr_file = read_airr_file(r.airr_file)
            csvs.append(airr_file)

            seq_plate_well_names = [
                wn[1] for wn in airr_file["sequence_id"].str.rsplit("_", n=1).to_list()
            ]
            nanobody_autonames = []
            for wn in seq_plate_well_names:
                try:
                    nanobody_autonames.append(
                        nanobody_autonames_lookup[
                            (
                                r.seq,
                                PlateLocations.labels.index(_extract_well(wn))
                                + 1
                                - r.well_pos_offset,
                            )
                        ]
                    )
                except ValueError:
                    nanobody_autonames.append("n/a (well unparseable)")
                except KeyError:
                    nanobody_autonames.append("n/a (index not found)")
            airr_file["nanobody_autoname"] = nanobody_autonames

        df = pd.concat(csvs)

        # Get number of matches per CDR3 sequence
        cdr3_counts = df["cdr3_aa"].value_counts()
        cdr3_counts.name = "cdr3_aa_count"
        df = df.merge(cdr3_counts, on="cdr3_aa", how="left")

        # Sort as required - Productive, number of CDR3 matches,
        # then CDR3 itself, then sequence ID
        df = df.sort_values(
            by=["productive", "cdr3_aa_count", "cdr3_aa", "sequence_id"],
            ascending=[False, False, True, True],
        )

        # Ensure we have the right columns in the right order
        df = df.loc[:, list(AIRR_IMPORTANT_COLUMNS) + ["nanobody_autoname"]]

        # Set index and replace NaN with None (null in JSON)
        df = df.replace({np.nan: None})

        # Indicator to show when cdr3 has changed from previous row
        df["new_cdr3"] = df["cdr3_aa"].shift(1).ne(df["cdr3_aa"])

        # Replace T/F with Y/N in productive and stop_codon columns
        for column in ("productive", "stop_codon"):
            df[column] = df[column].str.replace("F", "N").replace("T", "Y")

        sequences = (
            df["sequence_alignment_aa"].str.replace(".", "").replace("*", "X").tolist()
        )
        df["sequence"] = sequences
        df = df.drop("sequence_alignment_aa", axis=1)

        return JsonResponse({"records": df.to_dict(orient="records")})

    @action(
        detail=False,
        methods=["GET"],
        name="Search sequencing results by sequence.",
        url_path="searchcdr3/(?P<query>[A-Za-z]+)",
    )
    def search_sequencing_run_results(self, request, query):
        """Get sequencing run results by CDR3 sequence search."""
        srs = SequencingRunResults.objects.select_related("sequencing_run")
        results = []
        query = query.upper()
        for sr in srs:
            airr_file = read_airr_file(sr.airr_file)
            airr_file = airr_file[airr_file.cdr3_aa.notna()]
            if not airr_file.empty:
                airr_file = airr_file[airr_file.cdr3_aa.str.contains(query)]
            if not airr_file.empty:
                airr_file.insert(
                    loc=0, column="sequencing_run", value=sr.sequencing_run_id
                )
                results.append(airr_file)

        return JsonResponse(
            {"matches": pd.concat(results).to_dict(orient="records") if results else []}
        )

    @action(
        detail=True,
        methods=["GET"],
        name="BLAST sequencing run results vs DB.",
        url_path="blast",
    )
    def get_blast_sequencing_run(self, request, pk):
        """BLAST sequencing run vs database."""
        query_type = self.request.query_params.get("queryType", "full")
        if query_type not in ("full", "cdr3"):
            raise ValueError(f"Unknown queryType: {query_type}")
        blast_str = run_blastp(pk, query_type=query_type)
        if not blast_str:
            return JsonResponse({"hits": []}, status=status.HTTP_404_NOT_FOUND)

        # Read query AIRR files for CDRs
        results = SequencingRunResults.objects.filter(sequencing_run_id=int(pk))

        airr_df = pd.concat(read_airr_file(r.airr_file) for r in results)
        airr_df = airr_df.set_index("sequence_id")

        ALIGN_PERC_THRESHOLD = 90
        E_VALUE_THRESHOLD = 0.05

        # Parse the JSON and keep the important bits
        blast_result = json.loads(blast_str)
        res = []
        for blast_run in blast_result["BlastOutput2"]:
            run_res = blast_run["report"]["results"]["search"]
            for blast_hit_set in run_res["hits"]:
                subject_title = blast_hit_set["description"][0]["title"]
                query_title = run_res["query_title"]
                if query_type == "cdr3":
                    # TODO: Make more robust
                    query_cdr3 = query_title[6:]
                else:
                    query_cdr3 = airr_df.at[query_title, "cdr3_aa"]
                    if pd.isna(query_cdr3):
                        query_cdr3 = None
                if subject_title.strip() == query_title.strip():
                    continue
                hsps = blast_hit_set["hsps"][0]
                assert len(blast_hit_set["description"]) == 1
                for hsps in blast_hit_set["hsps"]:
                    align_perc = round(
                        (hsps["align_len"] / run_res["query_len"]) * 100, 2
                    )
                    if align_perc < ALIGN_PERC_THRESHOLD:
                        continue
                    e_value = hsps["evalue"]
                    if e_value > E_VALUE_THRESHOLD:
                        continue
                    res.append(
                        {
                            "query_title": query_title.strip(),
                            "query_cdr3": query_cdr3,
                            "subject_title": subject_title.strip(),
                            "submatch_no": hsps["num"],
                            "query_seq": hsps["qseq"],
                            "subject_seq": hsps["hseq"],
                            "midline": hsps["midline"],
                            "bit_score": hsps["bit_score"],
                            "e_value": e_value,
                            "align_len": hsps["align_len"],
                            "align_perc": align_perc,
                            "ident_perc": round(
                                (hsps["identity"] / hsps["align_len"]) * 100, 2
                            ),
                        }
                    )

        # Sort the results
        res = sorted(
            res,
            key=lambda r: (r["query_cdr3"] or "", -r["align_perc"], -r["ident_perc"]),
        )

        return JsonResponse({"hits": res})


class GlobalFastaView(APIView):
    def get(self, request, format=None):
        """Download entire database as .fasta file."""
        fasta_data = get_db_fasta()

        fasta_filename = (
            f"antigenapp_database_{datetime.datetime.now().isoformat()}.fasta"
        )
        response = FileResponse(
            fasta_data,
            as_attachment=True,
            content_type="text/x-fasta",
            filename=fasta_filename,
        )
        response["Content-Disposition"] = f'attachment; filename="{fasta_filename}"'
        return response
