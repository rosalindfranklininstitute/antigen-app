import collections.abc
import io
import os
from tempfile import NamedTemporaryFile
from wsgiref.util import FileWrapper

import numpy as np
import openpyxl
import pandas as pd
from django.core.files.storage import default_storage
from django.db.models import Prefetch
from django.http import FileResponse, Http404, HttpResponse, JsonResponse
from django.shortcuts import redirect
from rest_framework import status
from rest_framework.decorators import action
from rest_framework.serializers import (
    ModelSerializer,
    StringRelatedField,
    ValidationError,
)
from rest_framework.viewsets import ModelViewSet

from antigenapi.bioinformatics.blast import (
    parse_blast_results,
    run_blastp,
    run_blastp_seq_run,
)
from antigenapi.bioinformatics.imgt import (
    AIRR_IMPORTANT_COLUMNS,
    as_fasta_files,
    load_sequences,
    read_airr_file,
    run_vquest,
)
from antigenapi.models import (
    ElisaPlate,
    ElisaWell,
    Nanobody,
    PlateLocations,
    SequencingRun,
    SequencingRunResults,
)
from antigenapi.utils.helpers import extract_well, read_seqrun_results
from antigenapi.views.elisa import _wells_to_tsv
from antigenapi.views.mixins import AuditLogMixin, DeleteProtectionMixin


class SequencingRunResultSerializer(ModelSerializer):
    """A serializer for sequencing run results."""

    added_by = StringRelatedField()

    class Meta:  # noqa: D106
        model = SequencingRunResults
        fields = ("seq", "well_pos_offset", "added_by", "added_date")


class SequencingRunSerializer(ModelSerializer):
    """A serializer for sequencing runs."""

    added_by = StringRelatedField()
    sequencingrunresults_set = SequencingRunResultSerializer(many=True, required=False)

    def validate_wells(self, data):
        """Check JSONField for wells is valid."""
        for idx, well in enumerate(data):
            if len(well.keys()) != 3:
                raise ValidationError(f"Extraneous keys in well {idx}")
            if "elisa_well" not in well:
                raise ValidationError(f"Missing elisa_well in well {idx}")

            if "plate" not in well:
                raise ValidationError(f"Missing plate in well {idx}")
            if not isinstance(well["plate"], int):
                raise ValidationError(f"Well {idx}'s plate is not an integer")
            if well["plate"] != (idx // 96):
                raise ValidationError(
                    f"Well {idx}'s plate should be "
                    f"{(idx // 96)} (found: {well['plate']})"
                )

            if "location" not in well:
                raise ValidationError(f"Missing location in well {idx}")
            if not isinstance(well["location"], int):
                raise ValidationError(f"Well {idx}'s location is not an integer")
            if well["location"] < 1 or well["location"] > 96:
                raise ValidationError(
                    f"Well {idx}'s location must be between 1 and 96 inclusive"
                )

            if not isinstance(well["elisa_well"], collections.abc.Mapping):
                raise ValidationError(f"elisa_well in well {idx} is not an object")
            if len(well["elisa_well"].keys()) != 2:
                raise ValidationError(f"Extraneous keys in well {idx}'s elisa_well")
            if "plate" not in well["elisa_well"]:
                raise ValidationError(f"Missing plate in well {idx}'s elisa_well")
            if not isinstance(well["elisa_well"]["plate"], int):
                raise ValidationError(
                    f"Well {idx}'s elisa_well's plate is not an integer"
                )
            if "location" not in well["elisa_well"]:
                raise ValidationError(f"Missing location in well {idx}'s elisa_well")
            if not isinstance(well["elisa_well"]["location"], int):
                raise ValidationError(
                    f"Well {idx}'s elisa_well's location is not an integer"
                )
            if (
                well["elisa_well"]["location"] < 1
                or well["elisa_well"]["location"] > 96
            ):
                raise ValidationError(
                    f"Well {idx}'s elisa_well's location must be "
                    "between 1 and 96 inclusive"
                )

            # TODO: Check elisa_plate is valid plate
            # TODO: Check all locations are unique on a plate

        # Sort by plate and location
        data = sorted(data, key=lambda well: (well["plate"], well["location"]))

        return data

    def validate_plate_thresholds(self, data):
        """Check JSONField for plate_thresholds is valid."""
        seen_plate_ids = set()
        for idx, thr in enumerate(data):
            if len(thr.keys()) != 2:
                raise ValidationError(f"Extraneous keys in plate_threshold {idx}")
            if "optical_density_threshold" not in thr:
                raise ValidationError(
                    f"Optical density threshold missing in plate_threshold {idx}"
                )
            if not isinstance(thr["optical_density_threshold"], (int, float)):
                raise ValidationError(
                    f"Plate threshold {idx}'s optical_density_threshold "
                    "is not an integer or float"
                )
            if thr["optical_density_threshold"] < 0:
                raise ValidationError(
                    f"Plate threshold {idx}'s optical_density_threshold "
                    "must be at least 0"
                )
            if "elisa_plate" not in thr:
                raise ValidationError(f"Elisa plate missing in plate_threshold {idx}")
            if not isinstance(thr["elisa_plate"], int):
                raise ValidationError(
                    f"Plate threshold {idx}'s elisa_plate is not an integer"
                )
            if thr["elisa_plate"] in seen_plate_ids:
                raise ValidationError(
                    f"Duplicate elisa_plate {thr['elisa_plate']}"
                    f" in plate_threshold {idx}"
                )
            seen_plate_ids.add(thr["elisa_plate"])

        # One batch query to verify all referenced plates exist.
        if seen_plate_ids:
            existing_ids = set(
                ElisaPlate.objects.filter(pk__in=seen_plate_ids).values_list(
                    "pk", flat=True
                )
            )
            missing = seen_plate_ids - existing_ids
            if missing:
                raise ValidationError(f"ELISA plate(s) not found: {sorted(missing)}")

        return data

    def validate(self, data):
        """Object-level validation for SequencingRun."""
        # Resolve effective values for PATCH (partial) requests so cross-field
        # checks work correctly even when only one of the two fields is sent.
        if self.instance is not None:
            effective_wells = data.get("wells", self.instance.wells)
            effective_thresholds = data.get(
                "plate_thresholds", self.instance.plate_thresholds
            )
        else:
            effective_wells = data.get("wells", [])
            effective_thresholds = data.get("plate_thresholds", [])

        # Cross-field: every ELISA plate referenced by a well must have a
        # threshold entry.  Pure Python set comparison — no extra DB query.
        threshold_plate_ids = {thr["elisa_plate"] for thr in effective_thresholds}
        unlisted = {
            w["elisa_well"]["plate"] for w in effective_wells
        } - threshold_plate_ids
        if unlisted:
            raise ValidationError(
                f"Wells reference ELISA plate(s) with no threshold: {sorted(unlisted)}"
            )

        # Layout lock: prevent changes to layout fields once results are attached.
        # wells, plate_thresholds and fill_horizontal together define which ELISA
        # wells map to which positions in the sequencing plates.  Changing any of
        # them after results have been uploaded would silently mis-map results to
        # the wrong source wells.
        if self.instance is None or not self.instance.sequencingrunresults_set.exists():
            return data

        locked = {
            "wells": "wells",
            "plate_thresholds": "plate thresholds",
            "fill_horizontal": "fill direction",
        }
        changed = [
            label
            for field, label in locked.items()
            if field in data and data[field] != getattr(self.instance, field)
        ]
        if changed:
            raise ValidationError(
                "Cannot modify "
                + ", ".join(changed)
                + " on a SequencingRun that already has results attached."
            )
        return data

    class Meta:  # noqa: D106
        model = SequencingRun
        fields = "__all__"
        read_only_fields = ["added_by", "added_date"]


class SequencingRunShortSerializer(SequencingRunSerializer):
    """Sequencing run serializer which excludes plate/well information."""

    class Meta:  # noqa: D106
        model = SequencingRun
        exclude = ("plate_thresholds", "wells")
        read_only_fields = ["added_by", "added_date"]


class SequencingRunViewSet(AuditLogMixin, DeleteProtectionMixin, ModelViewSet):
    """A view set for sequencing runs."""

    queryset = (
        SequencingRun.objects.all()
        .order_by("-added_date")
        .select_related("added_by")
        .prefetch_related(
            Prefetch(
                "sequencingrunresults_set",
                queryset=SequencingRunResults.objects.select_related("added_by"),
            )
        )
    )
    serializer_class = SequencingRunSerializer

    def perform_create(self, serializer):  # noqa: D102
        serializer.save(added_by=self.request.user)

    @action(
        detail=True,
        methods=["GET"],
        name="Download sequencing run submission file (xlsx).",
        url_path="submissionfile/(?P<submission_idx>[0-9]+)",
    )
    def download_submission_xlsx(self, request, pk, submission_idx):
        """Download sequencing run submission file (xlsx)."""
        # Load the Excel template — one directory up from this file (views/)
        fn = os.path.join(
            os.path.dirname(os.path.dirname(os.path.realpath(__file__))),
            "files",
            "sequencing-submission-form-v1.xlsx",
        )
        oxl = openpyxl.load_workbook(fn)

        # Get first worksheet
        ws = oxl.worksheets[0]

        # Get sequencing run object
        sr = self.get_object()
        well_dict = {}
        elisa_plates_to_load = set()
        submission_idx = int(submission_idx)
        for well in sr.wells:
            if well["plate"] != submission_idx:
                continue
            well_dict[PlateLocations.labels[well["location"] - 1]] = well["elisa_well"]
            elisa_plates_to_load.add(well["elisa_well"]["plate"])

        if not well_dict:
            raise Http404

        elisa_wells = {
            (ew.plate_id, ew.location): ew
            for ew in ElisaWell.objects.filter(plate__pk__in=elisa_plates_to_load)
            .select_related("plate")
            .select_related("antigen")
        }

        # Make modifications
        for row in range(3, 100):  # 96 wells
            try:
                # Sample name
                well = well_dict[ws[f"A{row}"].value]
            except KeyError:
                continue

            elisa_well = elisa_wells[(well["plate"], well["location"])]
            ws[f"B{row}"] = (
                f"{elisa_well.antigen.short_name}_"
                f"{elisa_well.plate.pan_round_concentration:g}"
                f"{PlateLocations.labels[elisa_well.location - 1]}"
            )
            # Own primer name
            ws[f"D{row}"] = "PHD_SEQ_FWD"
            ws[f"F{row}"] = "Yes"

        # Save to temp file
        with NamedTemporaryFile() as tmp:
            oxl.save(tmp.name)
            response = HttpResponse(
                FileWrapper(tmp),
                content_type="application/vnd.openxmlformats-officedocument"
                ".spreadsheetml.sheet",
            )

        response["Content-Disposition"] = (
            f'attachment; filename="sequencing-submission-form-'
            f'sr{pk}_{submission_idx}.xlsx"'
        )
        return response

    @action(
        detail=True,
        methods=["GET"],
        name="Download sequencing run submission file (xlsx).",
        url_path="submissionfile/(?P<submission_idx>[0-9]+)/tsv",
    )
    def download_sequencing_plate_tsv(self, request, pk, submission_idx):
        """Download sequencing run plate layout as .tsv file."""
        try:
            sr = SequencingRun.objects.get(id=int(pk))
        except SequencingRun.DoesNotExist:
            raise Http404

        wells = {
            w["location"]: w for w in sr.wells if w["plate"] == int(submission_idx)
        }
        plate_ids = [w["elisa_well"]["plate"] for w in wells.values()]
        elisa_wells = {
            (ew.plate_id, ew.location): ew
            for ew in ElisaWell.objects.filter(plate_id__in=plate_ids)
        }

        well_dat = []
        for i in range(1, 97):
            try:
                well = wells[i]
            except KeyError:
                well_dat.append("")
                continue

            elisa_well = elisa_wells[
                (well["elisa_well"]["plate"], well["elisa_well"]["location"])
            ]

            well_dat.append(
                f"{elisa_well.plate_id}:"
                f"{PlateLocations.labels[elisa_well.location - 1]} "
                f"[{elisa_well.antigen}]"
            )

        output = _wells_to_tsv(well_dat)

        response = HttpResponse(output, content_type="text/tab-separated-values")

        fn = f"sequencing_run_{pk}_plate_{submission_idx}.tsv"
        response["Content-Disposition"] = f'attachment; filename="{fn}"'

        return response

    @action(
        detail=True,
        methods=["GET", "PUT"],
        name="Upload/download sequencing run results file (.zip).",
        url_path="resultsfile/(?P<submission_idx>[0-9]+)",
    )
    def sequencing_run_results(self, request, pk, submission_idx):
        """Upload sequencing run results file (.zip)."""
        if request.method == "GET":
            return self.download_sequencing_run_results(request, pk, submission_idx)

        results_file = request.data["file"]

        # TODO: Validate results file in more detail
        if not results_file.name.endswith(".zip"):
            raise ValidationError("file", "Results file should be a .zip file")

        # Validate the plate number and submission idx (seq) from the URL
        try:
            sr = SequencingRun.objects.get(pk=int(pk))
        except SequencingRun.DoesNotExist:
            raise ValidationError(
                f"Sequencing run {pk} does not exist to attach results"
            )

        # Store well positions
        wells_expected_list = [
            PlateLocations.labels[w["location"] - 1]
            for w in sr.wells
            if w["plate"] == int(submission_idx)
        ]
        if not wells_expected_list:
            raise ValidationError(
                f"Plate index {submission_idx} not found in sequencing run {pk}"
            )

        # Run bioinformatics using .zip file
        try:
            seq_data_fh = results_file.temporary_file_path()
        except AttributeError:
            seq_data_fh = results_file.file
        try:
            seq_data = load_sequences(seq_data_fh)
        except ValueError as e:
            raise ValidationError({"file": str(e)})
        if len(seq_data) != len(wells_expected_list):
            raise ValidationError(
                {
                    "file": f"Upload contains data for {len(seq_data)} "
                    f"wells, expected {len(wells_expected_list)}"
                }
            )
        # Validate wells expected vs wells supplied
        try:
            wells_supplied_list = [extract_well(w) for w in seq_data.keys()]
        except IndexError:
            raise ValidationError(
                {
                    "file": "Unable to parse well names. "
                    "Ensure all .seq filenames end with a well."
                }
            )

        # Check for duplicates in the supplied well names, and error if so
        wells_supplied = set(wells_supplied_list)
        if len(wells_supplied_list) != len(wells_supplied):
            seen = set()
            seen_twice = list(
                set(w for w in wells_supplied_list if w in seen or seen.add(w))
            )

            raise ValidationError(
                {"file": f"Duplicate wells found in supplied list: {seen_twice}"}
            )

        # Check for duplicates in the expected well names, and error if so
        wells_expected = set(wells_expected_list)
        if len(wells_expected_list) != len(wells_expected):
            raise ValueError(
                f"Duplicate wells found in expected list for seq run {pk} "
                f"idx {submission_idx}"
            )

        # Apply an offset in case the wells are shifted
        offset = 0

        wells_supplied_int = [PlateLocations.labels.index(w) for w in wells_supplied]
        wells_expected_int = [PlateLocations.labels.index(w) for w in wells_expected]
        if min(wells_supplied_int) > min(wells_expected_int):
            offset = min(wells_supplied_int) - min(wells_expected_int)
            wells_expected = set(
                [PlateLocations.labels[w + offset] for w in wells_expected_int]
            )

        if wells_expected - wells_supplied:
            raise ValidationError(
                {
                    "file": f"Expected well(s) {wells_expected - wells_supplied} "
                    f"were not found in upload (h.offset: {offset})"
                }
            )
        if wells_supplied - wells_expected:
            raise ValidationError(
                {
                    "file": f"Unexpected well(s) {wells_supplied - wells_expected} "
                    f"were found in upload (h.offset: {offset})"
                }
            )

        # Convert to FASTA in-memory (vquest api handles chunking to
        # max 50 seqs per file)
        fasta_file = as_fasta_files(seq_data, max_file_size=None)[0]

        # Submit to vquest
        vquest_results = run_vquest(fasta_file)

        parameters_file_data = vquest_results["Parameters.txt"]
        vquest_airr_data = vquest_results["vquest_airr.tsv"]

        base_filename = f"SequencingResults_{pk}_{submission_idx}"

        # Link to nanobodies table
        sequences = read_airr_file(
            io.BytesIO(vquest_airr_data.encode()), usecols=("sequence_alignment_aa",)
        )
        sequences = (
            sequences["sequence_alignment_aa"]
            .dropna()
            .str.replace(".", "")
            .replace("*", "X")
            .unique()
        )
        nanobodies = Nanobody.objects.filter(sequence__in=sequences)

        # Create SequencingRunResults object
        srr, _ = SequencingRunResults.objects.update_or_create(
            sequencing_run=SequencingRun.objects.get(pk=int(pk)),
            seq=submission_idx,
            defaults={
                "added_by": request.user,
                "seqres_file": results_file,
                "well_pos_offset": offset,
            },
        )
        srr.nanobodies.set(nanobodies)

        srr.airr_file.save(
            f"{base_filename}_vquestairr.tsv", io.StringIO(vquest_airr_data)
        )
        srr.parameters_file.save(
            f"{base_filename}_vquestparams.txt", io.StringIO(parameters_file_data)
        )

        return JsonResponse(
            SequencingRunSerializer(SequencingRun.objects.get(pk=int(pk))).data
        )

    def download_sequencing_run_results(self, request, pk, submission_idx):
        """Download sequencing run results file (.zip)."""
        try:
            sr = SequencingRunResults.objects.get(
                sequencing_run_id=int(pk), seq=int(submission_idx)
            )
        except SequencingRunResults.DoesNotExist:
            raise Http404

        if default_storage.__class__.__name__ == "S3Storage":
            return redirect(sr.seqres_file.url)

        filename = sr.seqres_file.path
        response = FileResponse(open(filename, "rb"))
        return response

    @action(
        detail=True,
        methods=["GET"],
        name="Download IMGT results file (.airr).",
        url_path="resultsfile/(?P<submission_idx>[0-9]+)/airr",
    )
    def download_sequencing_run_airr(self, request, pk, submission_idx):
        """Download sequencing run results file (.zip)."""
        try:
            sr = SequencingRunResults.objects.get(
                sequencing_run_id=int(pk), seq=int(submission_idx)
            )
        except SequencingRunResults.DoesNotExist:
            raise Http404

        if default_storage.__class__.__name__ == "S3Storage":
            return redirect(sr.airr_file.url)

        filename = sr.airr_file.path
        response = FileResponse(open(filename, "rb"))
        return response

    @action(
        detail=True,
        methods=["GET"],
        name="Get sequencing results.",
        url_path="results",
    )
    def get_sequencing_run_results(self, request, pk):
        """Get sequencing results."""
        df = read_seqrun_results(pk, usecols=AIRR_IMPORTANT_COLUMNS)

        if df.empty:
            return JsonResponse({"records": []})

        # Get number of matches per CDR3 sequence
        cdr3_counts = df["cdr3_aa"].value_counts()
        cdr3_counts.name = "cdr3_aa_count"
        df = df.merge(cdr3_counts, on="cdr3_aa", how="left")

        # Sort as required - Productive, number of CDR3 matches,
        # then CDR3 itself, then sequence ID
        df = df.sort_values(
            by=["productive", "cdr3_aa_count", "cdr3_aa", "sequence_id"],
            ascending=[False, False, True, True],
        )

        # Ensure we have the right columns in the right order
        df = df.loc[:, list(AIRR_IMPORTANT_COLUMNS) + ["nanobody_autoname"]]

        # Indicator to show when cdr3 has changed from previous row
        df["new_cdr3"] = df["cdr3_aa"].shift(1).ne(df["cdr3_aa"])

        # Replace T/F with Y/N in productive and stop_codon columns
        for column in ("productive", "stop_codon"):
            df[column] = df[column].str.replace("F", "N").replace("T", "Y")

        sequences = (
            df["sequence_alignment_aa"].str.replace(".", "").replace("*", "X").tolist()
        )
        df["sequence"] = sequences
        df = df.drop("sequence_alignment_aa", axis=1)

        # Replace NaN with None (null in JSON) - must be done AFTER all column
        # processing, as str operations on None/NaN values re-introduce NaN
        df = df.replace({np.nan: None})

        return JsonResponse({"records": df.to_dict(orient="records")})

    @action(
        detail=False,
        methods=["GET"],
        name="Search sequencing results by sequence.",
        url_path="searchseq/(?P<query>[A-Za-z]+)",
    )
    def search_sequencing_run_results(self, request, query):
        """Get sequencing run results by sequence search."""
        search_region = self.request.query_params.get("searchRegion", "full")

        df = pd.concat(
            [
                read_seqrun_results(
                    seqrun, usecols=("sequence_id", "sequence_alignment_aa", "cdr3_aa")
                )
                for seqrun in SequencingRun.objects.values_list("pk", flat=True)
            ]
        )

        if df.empty:
            return JsonResponse({"matches": []})

        query = query.upper()

        if search_region == "cdr3":
            df = df[df.cdr3_aa.notna()]
            df = df[df["cdr3_aa"].str.contains(query)]
        else:
            df = df[df.sequence_alignment_aa.notna()]
            df = df[df["sequence_alignment_aa"].str.contains(query)]

        df = df[
            ["sequencing_run", "nanobody_autoname", "sequence_alignment_aa", "cdr3_aa"]
        ]

        return JsonResponse({"matches": df.to_dict(orient="records")})

    @action(
        detail=False,
        methods=["GET"],
        name="BLAST custom query vs DB.",
        url_path="blastseq/(?P<query>[A-Za-z]+)",
    )
    def blast_sequencing_run_results(self, request, query):
        """BLAST sequencing run vs database."""
        search_region = self.request.query_params.get("searchRegion", "full")
        if search_region not in ("full", "cdr3"):
            raise ValueError(f"Unknown searchRegion: {search_region}")

        if len(query) > 1024:
            raise ValueError("Query string exceeds maximum length of 1024")

        query_str = f"> QuerySequence\n{query}\n"
        blast_str = run_blastp(
            query_str,
            query_type="cdr3_unagg" if search_region == "cdr3" else search_region,
        )

        if not blast_str:
            return JsonResponse({"hits": []}, status=status.HTTP_404_NOT_FOUND)

        return JsonResponse(
            {
                "hits": parse_blast_results(
                    blast_str,
                    search_region,
                    e_value_threshold=np.inf,
                    align_perc_theshold=0,
                )
            }
        )

    @action(
        detail=True,
        methods=["GET"],
        name="BLAST sequencing run results vs DB.",
        url_path="blast",
    )
    def get_blast_sequencing_run(self, request, pk):
        """BLAST sequencing run vs database."""
        query_type = self.request.query_params.get("queryType", "full")
        if query_type not in ("full", "cdr3"):
            raise ValueError(f"Unknown queryType: {query_type}")
        blast_str = run_blastp_seq_run(pk, query_type=query_type)
        if not blast_str:
            return JsonResponse({"hits": []}, status=status.HTTP_404_NOT_FOUND)

        # Read query AIRR files for CDRs
        airr_df = read_seqrun_results(pk, usecols=("sequence_id", "cdr3_aa"))
        airr_df = airr_df.set_index("nanobody_autoname")

        return JsonResponse(
            {"hits": parse_blast_results(blast_str, query_type, airr_df)}
        )
